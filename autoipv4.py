import os
import socket
import customtkinter
import tkinter
from tkinter import *
import tkinter.messagebox
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
from math import log2,ceil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import ipaddress
from customtkinter import CTk
from tkinter import simpledialog
import pyperclip
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import sys

def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
                # PyInstaller creates a temp folder and stores path in _MEIPASS
                base_path = sys._MEIPASS
        except Exception:
                base_path = os.path.abspath(".")
        
        return os.path.join(base_path, relative_path)




customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("dark-blue")  # Themes: "blue" (standard), "green", "dark-blue"


def change_scaling_event(app, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)
       
def change_appearance_mode_event(app, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)
       

@staticmethod
def save_feedback_to_excel(feedback_data):
       
        # Load the existing workbook
        wb = load_workbook(resource_path("feed_me.xlsx"))
       
        # Select the active sheet
        sheet = wb.active
       
        # Find the next available row to append feedback
        next_row = sheet.max_row + 1
       
        # Write the feedback data to the Excel file
        for row_num, feedback_row in enumerate(feedback_data, start=0):
            for col_num, value in enumerate(feedback_row, start=1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}{next_row+row_num}"] = value
       
        # Save the workbook
        wb.save(resource_path("feed_me.xlsx"))


def send_feedback_email(feedback):
        # Email configuration
        sender_email = "cyberwhizy@gmail.com" #update this with your own email address dedicated for the application.
        receiver_email = "cyberwhizy@gmail.com" #Update this with your own email address dedicated for the application.
        email_subject = "Feedback from Your App"
        smtp_server = "smtp.gmail.com"  # Update with your SMTP server details
        smtp_port = 587  # Update with your SMTP port

        # Create a multipart message to include both plain text and HTML
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = email_subject

        # Add the plain text part of the email
        text = f"Feedback received:\n{feedback}"
        plain_text = MIMEText(text, "plain")
        message.attach(plain_text)

        # Send the email using SMTP
        try:
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(sender_email, "neqi irpw bfgb oqov")  # Replace with your email password or use a secure method for authentication
                smtp_connection.sendmail(sender_email, receiver_email, message.as_string())
                smtp_connection.quit()
        except Exception as e:
                error_message = f"Failed to send feedback, make sure you have a stable internet connection: {e}"
                messagebox.showerror("Error", error_message)

       


def provide_feedback():
        feedback_dialog = customtkinter.CTkInputDialog(title="Feedback", text="Please provide your feedback")
        feedback = feedback_dialog.get_input()

        if feedback:
                feedback_data = [["User", feedback]]
                save_feedback_to_excel(feedback_data)
                send_feedback_email(feedback)  # Send feedback via email
                messagebox.showinfo("Feedback", "Thank you for your feedback!")
        else:
                messagebox.showwarning("Feedback", "No feedback provided.")

           
def open_input_dialog_event():
        dialog = customtkinter.CTkInputDialog(title="Feedback", text="Please provide your feedback")
        #print("Feedback:", dialog.get_input())


   


   
   


       
def show_help():
        # Display the help message in the messagebox
        messagebox.showinfo("Help", ("""""This is a help message for the Automatic IPv4 Calculator.
        You can provide an IP address and a CIDR value to calculate various information such as subnet mask, network address, broadcast address, number of subnets, hosts per subnet, network class, and reverse DNS lookup.

        Usage:
        For the SUBNET calculator
        1. Enter the IP address octets (e.g., 192.168.0.1) and the CIDR value (e.g., 24).
        2. Click the 'Calculate' button to perform the calculations.
        3. The results will be displayed in the corresponding fields.
        4. You can copy the results to the clipboard using the 'Copy Results' button.

        For the VLSM calculator 
        1. Enter the ip address/cidr in the provide field.
        2. Input the number of subnets required and press the select button.
        3. Fill in the hosts required in each subnet and then press calculate.
        4. The results are displayed below after the calculations are performed. 

        Note:
        - The input validation ensures that valid IP addresses and CIDR values are provided.
        - The 'Clear' button clears all input and output fields of the subnet calculator.
        - The select button also initiates another instance of calculations."""""))
       
       
   
 
       
       

 
       
       
   

    #THE METHODS FOR THE CALCULATE BUTTON TRIGGER ARE DEFINED BELOW
def calculate():
        # Get the input values
        oct1 = oct1_entry.get()
        oct2 = oct2_entry.get()
        oct3 = oct3_entry.get()
        oct4 = oct4_entry.get()
        cidr = cidr_entry.get()
        # Validate the input values
        if validate_input(oct1, oct2, oct3, oct4, cidr):
                ip = oct1 + '.' + oct2 + '.' + oct3 + '.' + oct4
                cidr = int(cidr)

                # Calculate the results
                subnet_mask = calculate_subnet_mask(cidr)
                network_address = calculate_network_address(ip, subnet_mask)
                broadcast_address = calculate_broadcast_address(ip, subnet_mask)
                num_subnets = calculate_number_of_subnets(cidr)
                hosts_per_subnet = calculate_hosts_per_subnet(cidr)
                network_class = calculate_network_class(ip)
                hostname = reverse_dns_lookup(ip)
                hosts_rane =calculate_hosts_range(ip, subnet_mask)
                total_hosts = calculate_total_hosts(cidr)
               


               

                # Update the output fields
                subnet_mask_entry.delete(0, customtkinter.END)
                subnet_mask_entry.insert(customtkinter.END, subnet_mask)
                network_address_entry.delete(0, customtkinter.END)
                network_address_entry.insert(customtkinter.END, network_address)
                broadcast_address_entry.delete(0, customtkinter.END)
                broadcast_address_entry.insert(customtkinter.END, broadcast_address)
                number_of_subnets_entry.delete(0, customtkinter.END)
                number_of_subnets_entry.insert(customtkinter.END, num_subnets)
                hosts_per_subnet_entry.delete(0, customtkinter.END)
                hosts_per_subnet_entry.insert(customtkinter.END, hosts_per_subnet)
                network_class_entry.delete(0, customtkinter.END)
                network_class_entry.insert(customtkinter.END, network_class)
               
               
                try:
                        hostname = reverse_dns_lookup(ip)  # Perform reverse DNS lookup
                except socket.herror:
                        hostname = "Hostname not found"

                reverse_dns_entry.delete(0, customtkinter.END)
                reverse_dns_entry.insert(customtkinter.END, hostname)

               
                hosts_range = calculate_hosts_range(ip, subnet_mask)
                network_class = calculate_network_class(ip)
                network_class_entry.delete(0, customtkinter.END)
                network_class_entry.insert(customtkinter.END, network_class)
                hosts_range_entry.delete(0, customtkinter.END)
                hosts_range_entry.insert(customtkinter.END, hosts_range)
                total_hosts_entry.delete(0, customtkinter.END)
                total_hosts_entry.insert(customtkinter.END, total_hosts)
               
       

               

               
               
               
                result_text = f"IP ADDRESS: {ip}/{cidr}\n"
                result_text += f"SUBNETMASK: {subnet_mask}\n"
                result_text += f"NETWORK ADDRESS: {network_address}\n"
                result_text += f"BROADCAST ADDRESS: {broadcast_address}\n"
                result_text += f"NUMBER OF SUBNETS: {num_subnets}\n"
                result_text += f"HOSTS PER SUBNET: {hosts_per_subnet}\n"
                result_text += f"NETWORK CLASS: {network_class}\n"
                result_text += f"REVERSE DNS LOOKUP: {hostname}\n"
                result_text += f"HOSTSRANGE: {hosts_range}\n"
                result_text += f"NUMBER OF HOSTS: {total_hosts}\n"
                result_text += f"Developed by Douglas"



               
                app.result_text = result_text
                messagebox.showinfo("Results", result_text)
               
        else:
                       
                messagebox.showerror("Error", "Please enter all the required fields.")
               
   
   
   
   
# The validate_input method is called by the calculate method to validate the input values entered by the user. It checks if the IP address and CIDR value are valid. If the input values are valid, it returns True. Otherwise, it returns False.
def validate_input( oct1_entry, oct2_entry, oct3_entry, oct4_entry, cidr):
        # Check if any entry field is empty and if empty, show an error message
        if not oct1_entry or not oct2_entry or not oct3_entry or not oct4_entry or not cidr:
            messagebox.showerror("Error", "All fields must be filled")
            return False                  

        # Check if the octets are valid integers between 1 and 255
        try:
            oct1_entry = int(oct1_entry)
            oct2_entry = int(oct2_entry)
            oct3_entry = int(oct3_entry)
            oct4_entry = int(oct4_entry)
        except ValueError:
            messagebox.showerror("Error", "Invalid IP address, please enter the right values")
            return False

        if not (0 <= oct1_entry <= 255) or not (0 <= oct2_entry <= 255) or not (0 <= oct3_entry <= 255) or not (0 <= oct4_entry <= 255):
            messagebox.showerror("Error", "Invalid IP address, please enter the right values")
            return False

        # Check if the CIDR value is a valid integer between 0 and 32
        try:
            cidr = int(cidr)
        except ValueError:
            messagebox.showerror("Error", "Invalid CIDR value")
            return False

        if not (0 <= cidr <= 32):
            messagebox.showerror("Error", "Invalid CIDR value")
            return False

        return True


           
def reset():
       
        # Reset all input and output fields
        oct1_entry.delete(0, customtkinter.END)
        oct2_entry.delete(0, customtkinter.END)
        oct3_entry.delete(0, customtkinter.END)
        oct4_entry.delete(0, customtkinter.END)
        cidr_entry.delete(0, customtkinter.END)
        subnet_mask_entry.delete(0, customtkinter.END)
        network_address_entry.delete(0, customtkinter.END)
        broadcast_address_entry.delete(0, customtkinter.END)
        number_of_subnets_entry.delete(0, customtkinter.END)
        hosts_per_subnet_entry.delete(0, customtkinter.END)
        network_class_entry.delete(0, customtkinter.END)
        reverse_dns_entry.delete(0, customtkinter.END)
        hosts_range_entry.delete(0, customtkinter.END)
        total_hosts_entry.delete(0, customtkinter.END)

@staticmethod
def calculate_subnet_mask(cidr):
        subnet_mask = []
        for i in range(4):
            if cidr >= 8:
                subnet_mask.append('255')
                cidr -= 8
            else:
                subnet_mask.append(str(256 - (2 ** (8 - cidr))))
                cidr = 0

        return '.'.join(subnet_mask)

@staticmethod
def calculate_network_address(ip, subnet_mask):
        octs = ip.split('.')
        mask_octets = subnet_mask.split('.')

        network_address = []
        for i in range(4):
            network_address.append(str(int(octs[i]) & int(mask_octets[i])))

        return '.'.join(network_address)

@staticmethod
def calculate_broadcast_address(ip, subnet_mask):
        octs = ip.split('.')
        mask_octets = subnet_mask.split('.')

        broadcast_address = []
        for i in range(4):
            broadcast_address.append(str(int(octs[i]) | (255 - int(mask_octets[i]))))

        return '.'.join(broadcast_address)

@staticmethod
def calculate_number_of_subnets(cidr):
        return str(2 ** (32 - cidr))

@staticmethod
def calculate_hosts_per_subnet(cidr):
        return str(2 ** (32 - cidr) - 2)

@staticmethod
def calculate_network_class(ip):
        octs = ip.split('.')
   
   
        if 1 <= int(octs[0]) <= 126:
            return "A"
        elif 128 <= int(octs[0]) <= 191:
            return "B"
        elif 192 <= int(octs[0]) <= 223:
            return "C"
        elif 224 <= int(octs[0]) <= 239:
            return "D"
        elif 240 <= int(octs[0]) <= 255:
            return "E"
        else:
            return ""

@staticmethod
def reverse_dns_lookup(ip):
        try:
            hostname = socket.gethostbyaddr(ip)[0]
            return hostname
        except socket.herror:
            return "Hostname not found"


@staticmethod
def calculate_total_hosts(cidr):
        """
        Calculate the total number of hosts in a subnet based on the CIDR notation.
        """
        host_bits = 32 - cidr
        return 2 ** host_bits - 2
   
   
@staticmethod
def calculate_hosts_range(ip, subnet_mask):
        """
        Calculate the host range (start and END IP addresses) for a given IP address and subnet mask.
        """
        ip_parts = ip.split('.')
        mask_parts = subnet_mask.split('.')

        network_address = []
        for i in range(4):
            network_address.append(str(int(ip_parts[i]) & int(mask_parts[i])))

        start_ip_parts = network_address.copy()
        start_ip_parts[-1] = str(int(start_ip_parts[-1]) + 1)
        start_ip = '.'.join(start_ip_parts)

        end_ip_parts = network_address.copy()
        end_ip_parts[-1] = str(int(end_ip_parts[-1]) + 254)
        end_ip = '.'.join(end_ip_parts)

        return  start_ip, end_ip



def copy_results():
# Get the result text from the result_text variable
        if hasattr(app, 'result_text'):
               
                
                pyperclip.copy(app.result_text)
                messagebox.showinfo("Success", "Results copied to clipboard")
        else:
               
                messagebox.showerror("Error", "No results to copy")
           

 
   

app=customtkinter.CTk()
app.title("Automatic IPv4 calculator")
app.iconbitmap(resource_path("appicon.ico"))
app.geometry(f"{960}x{580}")
app._windows_set_titlebar_color(color_mode="System")
app.resizable(False, False)

# Set a working directory to the script's location
os.chdir(os.path.dirname(os.path.abspath(__file__)))






# Create the title label widget
app.title_label = customtkinter.CTkLabel( master=app, text="AUTOMATIC IPv4 CALCULATOR", font=customtkinter.CTkFont(size=20, weight="bold"),)
app.title_label.grid(row=0, column=0,padx=10, pady=10, sticky="w", columnspan=20)
       
#Expand the first column to fill the window
app.grid_columnconfigure(0, weight=1)

       
#set the userappearance mode and scaling mode
#label for appearance mode
appearance_label = customtkinter.CTkLabel(master=app, text="Mode:")
appearance_label.grid(row=0, column=1, padx=10, pady=10, sticky="e")
appearance_mode_optionmenu = customtkinter.CTkOptionMenu(app,values=["Light", "Dark", "System"], command= lambda v:change_appearance_mode_event(app,v))
appearance_mode_optionmenu.grid(row=0,column=2, padx=10, pady=(10), sticky="e")
appearance_mode_optionmenu.set("Dark")
       

         
# Label for UI scaling mode
scaling_label = customtkinter.CTkLabel(master=app, text="UI Scaling:")
scaling_label.grid(row=0, column=3, padx=10, pady=10, sticky="e")
scaling_optionmenu = customtkinter.CTkOptionMenu(app,values=["80%", "90%", "100%", "110%", "120%"], command=lambda v:change_scaling_event(app,v))
scaling_optionmenu.grid(row=0, column=4, padx=10, pady=(10), sticky="e")
scaling_optionmenu.set("100")
       
       
       
       
#create a tabviewframe for the app that handles two tabs of SUBNET and VLSM input fields, labels and other widgets
tabview_frame = customtkinter.CTkTabview(master=app, width=950, height=580, corner_radius=0)
#tabview_frame.pack(fill="both", expand=True, padx=10, pady=10)
tabview_frame.place(x=1, y=50)
       
#add the two tabs of SUBNET and VLSM to the tabview a
tabview_frame.add("SUBNET CALCULATOR")
tabview_frame.add("VLSM CALCULATOR")
       
#configure the added tabs of SUBNET and VLSM
tabview_frame.tab("SUBNET CALCULATOR").grid_columnconfigure(0, weight=1,)

#configure the VLSM tab according to fit all the contents
tabview_frame.tab("VLSM CALCULATOR").grid_columnconfigure(0, weight=1,)
tabview_frame.tab("VLSM CALCULATOR").grid_columnconfigure(1, weight=1,)
tabview_frame.tab("VLSM CALCULATOR").grid_columnconfigure(2, weight=1,)
       

       
# Set the size of the tabview frames
tabview_frame.configure(width=950, height=550)


#create a scrollable frame with in the tabview frame
vlsm_main_frame = customtkinter.CTkScrollableFrame(tabview_frame.tab("VLSM CALCULATOR"),width=950, height=580, corner_radius=0)
vlsm_main_frame.grid(row=0, column=0,)

#create frames for the SUBNET CALCULATOR tab for labels of ip address throogh reverse dns lookup
frame1 = customtkinter.CTkFrame(master=tabview_frame.tab("SUBNET CALCULATOR"), width=180, height=500, corner_radius=0)
frame1.pack(fill="both",side="left", expand=True, padx=10, pady=10, )
frame1.place(x=0, y=0)
       
       
#create frame2 for the SUBNET CALCULATOR tab for input fields of ip address throogh reverse dns lookup
frame2 = customtkinter.CTkFrame(master=tabview_frame.tab("SUBNET CALCULATOR"), width=840, height=500, corner_radius=0)
frame2.pack(fill="both",side="right", expand=True, padx=10, pady=10)
frame2.place(x=180, y=0)
       
# Set the font and weight of the tab names      
label_font = customtkinter.CTkFont(size=14, weight="bold")
frame_width = frame1.winfo_width()


#create frames for the SUBNET CALCULATOR tab for labels of ip address throogh reverse dns lookup
frame1 = customtkinter.CTkFrame(master=tabview_frame.tab("SUBNET CALCULATOR"), width=180, height=500, corner_radius=0)
frame1.pack(fill="both",side="left", expand=True, padx=10, pady=10, )
frame1.place(x=0, y=0)
       
       
#create frame2 for the SUBNET CALCULATOR tab for input fields of ip address throogh reverse dns lookup
frame2 = customtkinter.CTkFrame(master=tabview_frame.tab("SUBNET CALCULATOR"), width=840, height=500, corner_radius=0)
frame2.pack(fill="both",side="right", expand=True, padx=10, pady=10)
frame2.place(x=180, y=0)
       
       
label_font = customtkinter.CTkFont(size=14, weight="bold")
frame_width =frame1.winfo_width()
       
       
#create labels for the SUBNET CALCULATOR tab for ip address throogh reverse dns lookup
label1 = customtkinter.CTkLabel(master=frame1, text="  IP  ADDRESS", font=label_font)
label1.pack(fill="both", expand=False, padx=150, pady=100,)
label1.place(x=10, y=10, )
label2 = customtkinter.CTkLabel(master=frame1, text="  SUBNET  MASK", font=label_font)
label2.pack(fill="both", expand=False, padx=150, pady=10)
label2.place(x=10, y=50, )
label3 = customtkinter.CTkLabel(master=frame1, text="NETWORK ADDRESS", font=label_font)
label3.pack(fill="both", expand=False, padx=150, pady=10)
label3.place(x=10, y=100)
label4 = customtkinter.CTkLabel(master=frame1, text="BROADCAST ADDRESS", font=label_font)
label4.pack(fill="both", expand=False, padx=150, pady=10)
label4.place(x=10, y=150)
label5 = customtkinter.CTkLabel(master=frame1, text="NUMBER OF SUBNETS", font=label_font)
label5.pack(fill="both", expand=False, padx=150, pady=10)
label5.place(x=10, y=200)
label6 = customtkinter.CTkLabel(master=frame1, text="HOSTS PER SUBNET", font=label_font)
label6.pack(fill="both", expand=False, padx=150, pady=10)
label6.place(x=10, y=250)
label7 = customtkinter.CTkLabel(master=frame1, text="NETWORK CLASS", font=label_font )
label7.pack(fill="both", expand=False, padx=150, pady=10)
label7.place(x=10, y=300)
label8 = customtkinter.CTkLabel(master=frame1, text="REVERSE DNS LOOKUP", font=label_font)
label8.pack(fill="both", expand=False, padx=150, pady=10)
label8.place(x=10, y=350)
label9 = customtkinter.CTkLabel(master=frame1, text="HOST RANGE", font=label_font)
label9.pack(fill="both", expand=False, padx=150, pady=10)
label9.place(x=10, y=400)
label10 = customtkinter.CTkLabel(master=frame1, text="TOTAL HOSTS", font=label_font)
label10.pack(fill="both", expand=False, padx=150, pady=10)
label10.place(x=10, y=450)
       
#create the entry fields for ip address and other parameters on frame2
#but first create the entry fields for ip address in octects and the /CDIR notation label and entry field
oct1_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font, placeholder_text="192",justify="center")
oct1_entry.grid(row=0, column=0, padx=10, pady=10)
oct2_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,placeholder_text="168", justify="center")
oct2_entry.grid(row=0, column=1, padx=10, pady=10)
oct3_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,placeholder_text="10",justify="center")
oct3_entry.grid(row=0, column=2, padx=10, pady=10)
oct4_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font, placeholder_text="100", justify="center")
oct4_entry.grid(row=0, column=3, padx=10, pady=10)

# create the /CDIR notation label and entry field
# first, create a / label
cidr_label = customtkinter.CTkLabel(master=frame2, text="/", font=label_font)
cidr_label.grid(row=0, column=4, padx=10, pady=10)

# next, create the cidr_entry field
cidr_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,placeholder_text="32", justify="center")
cidr_entry.grid(row=0, column=5, padx=10, pady=10)
       
#create entry fields for the subnet mask through reverse dns lookup
subnet_mask_entry = customtkinter.CTkEntry(master=frame2, width=100,font=label_font,justify="center")
subnet_mask_entry.grid(row=1, column=0, padx=10, pady=10,sticky="we", columnspan=6)
       
#set the gridcolumnconfigure of the frame2 to expand
frame2.grid_columnconfigure(0, weight=1)
       
#Add the rest of the entry fields
network_address_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
network_address_entry.grid(row=2, column=0, padx=10, pady=10, sticky="we", columnspan=6)
broadcast_address_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
broadcast_address_entry.grid(row=3, column=0, padx=10, pady=10,sticky="we", columnspan=6)
number_of_subnets_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
number_of_subnets_entry.grid(row=4, column=0, padx=10, pady=10, sticky="we", columnspan=6)
hosts_per_subnet_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
hosts_per_subnet_entry.grid(row=5, column=0, padx=10, pady=10,sticky="we", columnspan=6)
network_class_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
network_class_entry.grid(row=6, column=0, padx=10, pady=10, sticky="we", columnspan=6)
reverse_dns_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
reverse_dns_entry.grid(row=7, column=0, padx=10, pady=10, sticky="we", columnspan=6)
hosts_range_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
hosts_range_entry.grid(row=8, column=0, padx=10, pady=10, sticky="we", columnspan=6)
total_hosts_entry = customtkinter.CTkEntry(master=frame2, width=100, font=label_font,justify="center")
total_hosts_entry.grid(row=9, column=0, padx=10, pady=10, sticky="we", columnspan=6)
       
       
#create action buttons for the SUBNET calculator
calculate_button = customtkinter.CTkButton(master=frame2, text="CALCULATE", font=label_font, corner_radius=0, command=calculate)
calculate_button.grid(row=0, column=6, padx=(12), pady=(20,0), sticky="nswe", rowspan=2)
clear_button = customtkinter.CTkButton(master=frame2, text="CLEAR", font=label_font, corner_radius=0, command=reset)
clear_button.grid(row=2, column=6, padx=12, pady=(20,0), sticky="nswe",  rowspan=2)
copy_button = customtkinter.CTkButton(master=frame2, text="COPY", font=label_font, corner_radius=0, command=copy_results)
copy_button.grid(row=4, column=6, padx=12, pady=(20,0), sticky="nswe",  rowspan=2)
help_button = customtkinter.CTkButton(master=frame2, text="HELP", font=label_font, corner_radius=0, command=show_help)
help_button.grid(row=6, column=6, padx=12, pady=(20,0), sticky="nswe",  rowspan=2)
feedback_button = customtkinter.CTkButton(master=frame2, text="FEEDBACK", font=label_font, corner_radius=0, command=provide_feedback)
feedback_button.grid(row=8, column=6, padx=12, pady=(20, 0), sticky="nswe", rowspan=2)




#       """""""""""""""""""""""""""""""VLSM CALCTULATOR TAB METHODS START FROM THIS LINE"""""""""""""""""""""""""""""""""""""""""""



# first list all the data
subnetdata = []
tree = None
yscrollbar = None
xscrollbar = None
 
def update_entry():
        global subnets
        s = int(subnets.get())
        #first lets clear the hosts_frame
        for widget in hosts_frame.winfo_children():
            widget.destroy()
           
           
        #clear the results_frame
        for widget in results_frame.winfo_children():
            widget.destroy()
        subnetdata.clear()
       
        #craete headers for the entries
        index_header = customtkinter.CTkLabel(master=vlsm_frame, text="#",font=label_font, width=50)
        index_header.grid(row=2, column=0, padx=10, pady=10,sticky="w")
        subnet_header = customtkinter.CTkLabel(master=vlsm_frame, text="Subnet",font=label_font, width=50)
        subnet_header.grid(row=2, column=1, padx=10, pady=10,sticky="w")
        hosts_header = customtkinter.CTkLabel(master=vlsm_frame, text="Hosts",font=label_font, width=50)
        hosts_header.grid(row=2, column=2, padx=10, pady=10,sticky="w")
       
        for i in range(s):
           
            #create the index label
            index_label = customtkinter.CTkLabel(master=hosts_frame, text=i+1)
            index_label.grid(row=0+i, column=0, padx=10, pady=10,sticky="ew")
           
            #create the looping subnet entry
            subnet_entry = customtkinter.CTkEntry(master=hosts_frame, width=300, font=label_font, justify="center")
            subnet_entry.grid(row=0+i,column=1, padx=10, pady=10,sticky="w")
            subnet_entry.insert(customtkinter.END, f"Subnet {i}")
           
            #create the looping hosts entry
            hosts_entry = customtkinter.CTkEntry(master=hosts_frame, width=300, font=label_font, justify="center")
            hosts_entry.grid(row=0+i,column=2, padx=10, pady=10,sticky="w")
           
           
            subnetdata.append((subnet_entry, hosts_entry))
            #subnetdata.append((index_label, subnet_entry, hosts_entry))
        #create a calculate button
        vlsm_calculate_button = customtkinter.CTkButton(master=hosts_frame, text="CALCULATE", corner_radius=0, font=label_font, command=update_subnets)
        vlsm_calculate_button.grid(row=0+s, column=2)
       
       
       
       



def update_subnets():
        s = int(subnets.get())
        global tree, yscrollbar, xscrollbar
        # clear the results frame    
        for widget in results_frame.winfo_children():
                widget.destroy()
               
        # first lets get the data from the entries
        entry_data = calculate_vlsm(s)
       
        if tree is not None:
            tree.destroy()
        if yscrollbar is not None:
            yscrollbar.destroy()
        if xscrollbar is not None:
            xscrollbar.destroy()

       
        # create a treeview widget
        tree = ttk.Treeview(results_frame, show="headings", height=300)
        tree["columns"] = entry_data.columns.tolist()
       
        tree.selection_set(tree.get_children())  # Select all items by default

       
        # configure the column headings
        for i, column in enumerate(entry_data.columns):
           
            tree.heading("#" + str(i+1), text=column)
            tree.column("#" + str(i+1), width=100)
           
        # insert the data into the treeview
        for i in range(s):
            tree.insert("", "end", text=str(i+1), values=entry_data.iloc[i].tolist())
           
        # create a vertical scrollbar
       
        yscrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=yscrollbar.set)
        yscrollbar.grid(row=0, column=3, sticky="ns")
       
        # create a horizontal scrollbar
        xscrollbar = ttk.Scrollbar(results_frame, orient="horizontal", command=tree.xview)
        tree.configure(xscrollcommand=xscrollbar.set)
        xscrollbar.grid(row=1, column=0, sticky="ew")
       
        tree.grid(row=0, column=0, sticky="nsew")
       
        #configure the treeview
        #tree.columnconfigure(0, weight=1)
        #tree.columnconfigure(1, weight=1)
        #tree.columnconfigure(2, weight=1)
       
        # Show message box with the results
        results_info = entry_data.to_string()
        messagebox.showinfo("Results", results_info)

        
       
       
       
       
       
       
#create a function to calculate the subnets and hosts

def calculate_vlsm(s):
        #s=int(subnets.get())
        #first validate the ip/cidr entry
        try:
                network_address,prefixlen = validate_ipv4network()
        except:
                messagebox.showerror("ERROR", "INVALID IP ADDRESS/CIDR")
        assert validate_ipv4_address(network_address)
        columnheaders = ["Subnet","Total IPs","Hosts Needed", "Available Hosts","Unused Hosts", "Network Address", "CIDR", "Subnet Mask", "Usable Range", "Broadcast Address", "Wildcard Mask"]
        subnetsframe = pd.DataFrame(columns = columnheaders, index=range(s))
        reservedcount = 2
       
        for i in range(s):
                subnetsframe["Subnet"][i] = subnetdata[i][0].get()
                try:
                        subnetsframe["Hosts Needed"][i] = int(subnetdata[i][1].get())
                except:
                        messagebox.showerror("ERROR", "Invalid number of hosts,using 30 instead ")
                        subnetsframe["Hosts Needed"][i] = 30
                hostsbits = ceil(log2(subnetsframe["Hosts Needed"][i]+2))
                subnetsframe["Total IPs"][i] = 2**hostsbits
                subnetsframe["Available Hosts"][i] = subnetsframe["Total IPs"][i]-reservedcount
                subnetsframe["Unused Hosts"][i] = subnetsframe["Available Hosts"][i]-subnetsframe["Hosts Needed"][i]
                subnetsframe["CIDR"][i] = 32-hostsbits
       
        totalhosts = sum(subnetsframe["Total IPs"])
        maxprefixlen = 32-ceil(log2(totalhosts))
        if(prefixlen>maxprefixlen):
                messagebox.showinfo("NOTICE THIS", f"No CIDR provided, using/{maxprefixlen} instead")
                prefixlen = maxprefixlen
       
        addressdata = [0]*4
        octect = int(prefixlen/8)
       
        for i in range(octect):
                addressdata[i] = network_address[i]
               
        try:
                ipAddress = ipaddress.ip_address('.'.join(map(str,addressdata)))
        except:
                messagebox.showerror("ERROR", "Invalid ip addres, please enter a correct ipv4 address")
               
        for i in range(s):
                subnet = ipaddress.IPv4Network(str(ipAddress)+f"/{subnetsframe.CIDR[i]}", False)
               
                subnetsframe["Network Address"][i] = subnet.network_address
                subnetsframe["Subnet Mask"][i] = subnet.netmask
                subnetsframe["Broadcast Address"][i] = subnet.broadcast_address
                subnetsframe["Wildcard Mask"][i] = subnet.hostmask
               
                hostslist = list(subnet.hosts())
                subnetsframe["Usable Range"][i] = f"{(hostslist[0])} - {hostslist[-1]}"
                ipAddress += subnetsframe["Total IPs"][i]
         
        #print(subnetsframe.head())      
        return subnetsframe







def copy_results_to_clipboard():
        s = int(subnets.get())
        


        if not subnetdata:

                messagebox.showinfo("No Results", "Please calculate the results first.")
                return

        entry_data = calculate_vlsm(s)
        # Get the results from the entry_data DataFrame
        results_info = entry_data.to_string()
        
        try:

                # Copy the results to the clipboard
                pyperclip.copy(results_info)
                messagebox.showinfo("Success", "Results copied to clipboard!")
        except:

                messagebox.showerror("Failure", "Failed to copy results to clipboard.")
       

       
def validate_ipv4network():
        network_str = ip_cidr_entry.get()
        network_split = network_str.split("/")
        ip_address_str = network_split[0].split(".")
        ipvlsm = list(map(int,ip_address_str))
        if (len(network_split)!=2):
                messagebox.showerror("ERROR", "No cidr added")
                messagebox.showinfo("INFORMATION", "Using an auto cidr of 32")

                return ipvlsm,32
        prefixlen = int(network_split[1])
        return ipvlsm, prefixlen
def validate_network_classvlsm(ipvlsm):
        classdictionary = {0:'A',1:'B',2:'C',3:'D',4:'E'}
        classint = -1
        for val in ipvlsm:
                if val>0:
                        classint = -1
        if(ipvlsm[0]>=0):
                classint = max(classint,0)
        elif(ipvlsm[0]>=128):
                classint = max(classint,1)
        elif(ipvlsm[0]>=192):
                classint = max(classint,2)
        elif(ipvlsm[0]>=224):
                classint = max(classint,3)
        elif(ipvlsm[0]>=240):
                classint = max(classint,4)  
        return classint


def validate_ipv4_address(ipvlsm):
        try:
                if(len(ipvlsm) == 4):
                        return True
                else:
                        return False
               
        except:
                return
       
       
                                            
               

#         """"""""""""""WIDGETS FOR THE VLSM CALCULATOR"""""""""""""""""""          
       
           
# first create a frame for the vlsm
vlsm_frame = customtkinter.CTkFrame(vlsm_main_frame, height = 100, width = 950) #width=950, height=500)
vlsm_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nswe")
#vlsm_frame.pack(fill="both", expand=True, padx=10, pady=10)
#vlsm_frame.place(x=0,y=0)

# expand the first column to fill the entire frame
vlsm_frame.columnconfigure(0, weight=1)
vlsm_frame.columnconfigure(1, weight=1)
vlsm_frame.columnconfigure(2, weight=1)

# add widgets in the vlsm_frame
# add the ip/cidr label
ip_cidr_label = customtkinter.CTkLabel(master=vlsm_frame, text="IP ADDRESS/CIDR", font=label_font, width=50)
ip_cidr_label.grid(row=0, column=0, padx=10, pady=10, sticky="ew",)


# add the ip/cidr entry
ip_cidr_entry = customtkinter.CTkEntry(master=vlsm_frame, width=300, font=label_font, justify="center")
ip_cidr_entry.grid(row=0, column=1, padx=10, pady=10, sticky="w", )
ip_cidr_entry.insert(customtkinter.END, "192.168.4.0/24")

# add the subnets label
subnets_label = customtkinter.CTkLabel(master=vlsm_frame, text="SUBNETS", font=label_font, width=50)
subnets_label.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

# add the subnet_spinbox
subnets = Spinbox(master=vlsm_frame, from_=1, to=255, state="normal")
subnets.grid(row=1, column=1, padx=10, pady=10, sticky="w", )
subnets.config(width=20, font=label_font, background="#1C82AD")

# create a select button to update the entry hosts entry frame
select_button = customtkinter.CTkButton(master=vlsm_frame, text="SELECT", corner_radius=0, font=label_font, command=update_entry)
select_button.grid(row=0, column=2, padx=12, pady=(20,0),sticky="ns")

# Create the help button 
vlsm_help_button = customtkinter.CTkButton(master=vlsm_frame, text="HELP", corner_radius=0, font=label_font, command=show_help)
vlsm_help_button.grid(row=0, column=3, padx=12, pady=(20,0), sticky="ns")

# Create the feedback button 
vlsm_feedback_button = customtkinter.CTkButton(master=vlsm_frame, text="FEEDBACK", corner_radius=0, font=label_font, command=provide_feedback)
vlsm_feedback_button.grid(row=1, column=3, padx=12, pady=(20,0), sticky="ns")



# Create a button to copy results to clipboard
copy_button_vlsm = customtkinter.CTkButton(master=vlsm_frame, text="COPY", font=label_font,corner_radius=0, command=copy_results_to_clipboard)
copy_button_vlsm.grid(row=1, column=2, padx=12, pady=(20,0), sticky="ns")
       
       
       
# Entry for hosts in the subnet frame
hosts_frame = customtkinter.CTkFrame(vlsm_main_frame, height = 100, width = 950)#width=950,height=100)
hosts_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nswe")

# configure the hosts frame
hosts_frame.columnconfigure(0, weight=1)
hosts_frame.columnconfigure(1, weight=1)
hosts_frame.columnconfigure(2, weight=1)


       
# Create the results_frame.
results_frame = customtkinter.CTkFrame(vlsm_main_frame, height = 300, width = 950)
results_frame.grid(row=3,column=0, pady=10, sticky="nswe")



# configure the results frame coulmns
results_frame.columnconfigure(0, weight=1)
results_frame.columnconfigure(1, weight=1)
results_frame.columnconfigure(2, weight=1)

# configure the resulsts frame rows
results_frame.rowconfigure(0, weight=1)
results_frame.rowconfigure(1, weight=1)
results_frame.rowconfigure(2, weight=1)


       



 
   
app.mainloop()
